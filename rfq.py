import streamlit as st
import pandas as pd
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import tempfile
import os

st.set_page_config(page_title="Warehouse RFQ Generator", page_icon="🏭", layout="wide")

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
CATEGORY_OPTIONS = [
    "Storage System",
    "Material Handling",
    "Automated Storage System",
    "Dock Leveller",
    "Storage Container",
]

STORAGE_SYSTEM_ITEMS = [
    "Heavy-duty Racks", "Pallet Racking Systems", "Industrial Shelving",
    "Cantilever Racks", "Mezzanine Floors", "Tabular Racks", "Mobile Storage Racks",
]
MATERIAL_HANDLING_ITEMS = [
    "Forklifts", "Hand Pallet Trucks", "Electric Pallet Trucks", "Stackers",
    "Trolleys", "Conveyor Systems", "Scissor Lifts",
]
AUTOMATED_STORAGE_ITEMS = [
    "Vertical Carousel System", "Horizontal Carousel System",
]
DOCK_LEVELLER_ITEMS = [
    "Dock Levellers", "Dock Plates", "Loading Ramps",
]
STORAGE_CONTAINER_ITEMS = [
    "Plastic Bins", "Crates", "Pallets (Wood)", "Pallets (Plastic)",
    "Pallets (Metal)", "Storage Boxes",
]

CATEGORY_ITEMS = {
    "Storage System": STORAGE_SYSTEM_ITEMS,
    "Material Handling": MATERIAL_HANDLING_ITEMS,
    "Automated Storage System": AUTOMATED_STORAGE_ITEMS,
    "Dock Leveller": DOCK_LEVELLER_ITEMS,
    "Storage Container": STORAGE_CONTAINER_ITEMS,
}

UNIT_OPTIONS = ["", "Nos", "Pieces", "Sets", "Meters", "Sq.Ft", "Sq.M",
                "Kg", "Tons", "Liters", "Boxes", "Rolls", "Pairs", "Lots"]

# Automated Storage System spec rows (from Excel template)
AUTOMATED_MODEL_DETAILS = [
    (1, "Dimensions of VStore", "Height (mm)", "mm", "28000"),
    ("", "", "Width (mm)", "mm", "3200"),
    ("", "", "Depth (mm)", "mm", "3400"),
    ("", "", "Floor area (m2)", "m2", ""),
    ("", "", "1st Access Point Height (mm)", "mm", "836"),
    ("", "", "2nd Access Point Height (mm)", "mm", "5836"),
    ("", "", "3rd Access Point Height (mm)", "mm", "8836"),
    ("", "", "4th Access Point Height (mm)", "mm", "11836"),
    ("", "", "Dead weight of Machine (Kg)", "Kg", ""),
    ("", "", "Total Weight of Tray (Kg)", "Kg", ""),
    ("", "", "Total Weight of Machine (Kg)", "Kg", ""),
    ("", "", "Storage capacity (Kg)", "Kg", ""),
    ("", "", "Total Machine carrying capacity", "", ""),
    ("", "", "Total full weight (Kg)", "Kg", ""),
    (2, "Floor load", "Total (Kgs/sqm)", "Kgs/sqm", ""),
    (3, "Tray Details", "Usable width (mm)", "mm", ""),
    ("", "", "Usable depth (mm)", "mm", ""),
    ("", "", "Empty Tray weight", "Kg", ""),
    ("", "", "Area of each Trays (mm)", "mm", ""),
    ("", "", "Maximum load capacity (Kg)", "Kg", "465"),
    ("", "", "Number of Trays (Nos.)", "Nos", ""),
    ("", "", "Total area of all Trays (m2)", "m2", ""),
    (4, "Access time", "Maximum (Sec.)", "Sec", ""),
    ("", "", "Average (Sec.)", "Sec", ""),
    (5, "No Trays can Fetch", "No trays / Hour", "Nos/Hr", ""),
    (6, "Power supply", "", "", ""),
    (7, "Maximum Power rating", "", "", ""),
    (8, "Control Panel", "VStore standard control panel", "", ""),
    (9, "Height optimisation", "Provided for storage", "", ""),
    (10, "Operator panel", "", "", ""),
    (11, "Accessories", "Emergency stop", "", ""),
    ("", "", "Accident protection light curtains", "", ""),
    ("", "", "Lighting in the accessing area", "", ""),
]
AUTOMATED_KEY_FEATURES = [
    "Material Tracking", "Tray Details", "Inventory List", "Tray Call History",
    "Alarm History", "Item Code Search", "Bar Code Search", "Pick from BOM",
    "BOM Items List", "User Management, with backup and restore options",
]
AUTOMATED_INBUILT = [
    "Ergonomic tray positioning", "Variable frequency drives",
    "Tray uneven positioning sensor",
    "Light barrier for sensing material and operator intervention",
    "Operator panel with IPC", "Weight management system for sensing tray overload",
    "Tray Block option for Multiple users", "Password authentication",
    "Tray guide rail @ 50 pitch", "Total machine capacity 60 tone",
    "Expansion at later stage is possible", "Inventory management software",
]
AUTOMATED_INSTALLATION = [
    "Inventory Management Suite (IPC).",
    "Packing, Freight & Transit Insurance.",
    "Installation & Commissioning.", "Training.", "Warranty Period",
    "Unloading of material", "Material handling during the installation",
    "Power cable cost main junction Box to Vstore",
    "Biometric Access, Barcode Scanner", "MS office.",
    "Software Customization",
    "Machine Integration with ERP system will extra at Actual.",
    "UPS and Stabilizer with accessories Installation",
    "Equipment Movement & Installation location",
    "PEB Cladding and Civil Floor for outside installation",
]

# ─────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():   return Font(name="Arial", bold=True, size=10)
def normal_font():   return Font(name="Arial", size=10)
def title_font():    return Font(name="Arial", bold=True, size=12)
def section_font():  return Font(name="Arial", bold=True, size=11)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")   # light blue
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")  # softer blue
TITLE_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy (white text)

def style_header_cell(cell, fill=True):
    cell.font = header_font()
    cell.alignment = center_align()
    cell.border = thin_border()
    if fill:
        cell.fill = HEADER_FILL

def style_data_cell(cell, align="center"):
    cell.font = normal_font()
    cell.alignment = center_align() if align == "center" else left_align()
    cell.border = thin_border()


# ─────────────────────────────────────────────
#  EXCEL EXPORT — Standard groups
#  (Storage System / Material Handling / Dock Leveller)
# ─────────────────────────────────────────────
def export_standard_excel(category, df, project_name, company):
    wb = Workbook()
    ws = wb.active
    ws.title = category[:31]
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9   # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Title block ──
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REQUEST FOR QUOTATION"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.alignment = center_align()
    t.fill = TITLE_FILL

    ws.merge_cells("A2:F2")
    t2 = ws["A2"]
    t2.value = f"Category: {category}  |  Project: {project_name}  |  Company: {company}"
    t2.font = Font(name="Arial", bold=True, size=10)
    t2.alignment = center_align()
    t2.fill = SECTION_FILL
    t2.border = thin_border()

    # ── Column headers (row 4) ──
    headers = ["Sr.no", "Item Name", "Description / Specification", "Qty", "Unit", "Remarks"]
    col_widths = [8, 30, 55, 10, 12, 30]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=i, value=h)
        style_header_cell(c)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 22

    # ── Data rows ──
    for r_idx, (_, row) in enumerate(df.iterrows(), 5):
        item = str(row.get("Item Name", "")).strip()
        if not item:
            continue
        ws.row_dimensions[r_idx].height = 20
        vals = [
            r_idx - 4,
            item,
            str(row.get("Description / Specification", "")),
            str(row.get("Quantity", "")),
            str(row.get("Unit", "")),
            str(row.get("Remarks", "")),
        ]
        aligns = ["center", "left", "left", "center", "center", "left"]
        for c_idx, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            style_data_cell(c, a)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  EXCEL EXPORT — Automated Storage System
#  (exact Excel template structure)
# ─────────────────────────────────────────────
def export_automated_excel(df_model, model_header, df_kf, df_ib, df_ia,
                            project_name, company):
    wb = Workbook()
    ws = wb.active
    ws.title = "Automated Storage System"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # column widths matching template
    col_widths = {"A": 8, "B": 8, "C": 35, "D": 42, "E": 20, "F": 28}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    def mb(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def hdr(row, col, val, span_end_col=None, fill=HEADER_FILL, font_size=10):
        if span_end_col:
            mb(row, col, row, span_end_col)
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=font_size,
                      color="FFFFFF" if fill == TITLE_FILL else "000000")
        c.alignment = center_align()
        c.fill = fill
        c.border = thin_border()
        return c

    # ── Page title ──
    hdr(1, 1, "REQUEST FOR QUOTATION", span_end_col=6, fill=TITLE_FILL, font_size=13)
    ws.row_dimensions[1].height = 28
    mb(2, 1, 2, 6)
    c = ws.cell(row=2, column=1,
                value=f"Automated Storage System  |  Project: {project_name}  |  Company: {company}")
    c.font = Font(name="Arial", bold=True, size=10)
    c.alignment = center_align()
    c.fill = SECTION_FILL
    c.border = thin_border()
    ws.row_dimensions[2].height = 20

    row = 4  # start writing

    # ══ MODEL DETAILS ══
    mb(row, 1, row, 6)
    c = ws.cell(row=row, column=1, value="Model Details")
    c.font = section_font(); c.alignment = center_align()
    c.border = thin_border(); c.fill = SECTION_FILL
    ws.row_dimensions[row].height = 20
    row += 1

    mb(row, 1, row, 6)
    c = ws.cell(row=row, column=1, value=model_header)
    c.font = Font(name="Arial", bold=True, size=10); c.alignment = center_align()
    c.border = medium_border(); c.fill = HEADER_FILL
    ws.row_dimensions[row].height = 18
    row += 1

    # Model Details header row
    md_hdrs = ["Sr.no", "Category", "Description", "UNIT", "Requirement", "Vendor Reply"]
    for ci, h in enumerate(md_hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        style_header_cell(c)
    ws.row_dimensions[row].height = 22
    row += 1

    # group rows with merged Sr/Category
    # Track merge ranges for Sr and Category
    pending_sr_start = None
    pending_cat_start = None
    pending_sr_val = None
    pending_cat_val = None

    def flush_merges(end_row):
        nonlocal pending_sr_start, pending_cat_start, pending_sr_val, pending_cat_val
        if pending_sr_start and end_row > pending_sr_start:
            if end_row - 1 > pending_sr_start:
                mb(pending_sr_start, 1, end_row - 1, 1)
        if pending_cat_start and end_row > pending_cat_start:
            if end_row - 1 > pending_cat_start:
                mb(pending_cat_start, 2, end_row - 1, 2)
        pending_sr_start = pending_cat_start = None
        pending_sr_val = pending_cat_val = None

    for (sr, cat, desc, unit, req) in AUTOMATED_MODEL_DETAILS:
        ws.row_dimensions[row].height = 18
        if sr != "":
            flush_merges(row)
            pending_sr_start = row
            pending_sr_val = sr
            pending_cat_start = row
            pending_cat_val = cat
            c_sr = ws.cell(row=row, column=1, value=sr)
            c_sr.font = normal_font(); c_sr.alignment = center_align(); c_sr.border = thin_border()
            c_cat = ws.cell(row=row, column=2, value=cat)
            c_cat.font = Font(name="Arial", bold=True, size=10)
            c_cat.alignment = left_align(); c_cat.border = thin_border()
        else:
            c_sr = ws.cell(row=row, column=1, value="")
            c_sr.border = thin_border()
            c_cat = ws.cell(row=row, column=2, value="")
            c_cat.border = thin_border()

        for ci, v in enumerate([desc, unit, req, ""], 3):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = normal_font()
            c.alignment = center_align() if ci >= 4 else left_align()
            c.border = thin_border()
        row += 1

    flush_merges(row)
    row += 2

    # ══ KEY FEATURES ══
    mb(row, 1, row, 6)
    c = ws.cell(row=row, column=1, value="Key Features")
    c.font = section_font(); c.alignment = center_align()
    c.border = medium_border(); c.fill = SECTION_FILL
    ws.row_dimensions[row].height = 20
    row += 1

    for ci, h in enumerate(["Sr.no", "Description", "Status", "Remarks", "", ""], 1):
        if h:
            c = ws.cell(row=row, column=ci, value=h)
            style_header_cell(c)
        else:
            ws.cell(row=row, column=ci).border = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    for i, feat in enumerate(AUTOMATED_KEY_FEATURES, 1):
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=1, value=i).border = thin_border()
        ws.cell(row=row, column=1).font = normal_font()
        ws.cell(row=row, column=1).alignment = center_align()
        c2 = ws.cell(row=row, column=2, value=feat)
        c2.font = normal_font(); c2.alignment = left_align(); c2.border = thin_border()
        # Status + Remarks columns
        for ci in [3, 4]:
            kf_row = df_kf[df_kf["Description"] == feat] if df_kf is not None else pd.DataFrame()
            v = ""
            if not kf_row.empty:
                v = str(kf_row.iloc[0].get("Status" if ci == 3 else "Remarks", ""))
            c = ws.cell(row=row, column=ci, value=v)
            c.font = normal_font(); c.alignment = left_align(); c.border = thin_border()
        for ci in [5, 6]:
            ws.cell(row=row, column=ci).border = thin_border()
        if i == 1:
            remark = "All these key features including in Vendor Dashboard."
            c_rem = ws.cell(row=row, column=4, value=remark)
            c_rem.font = normal_font(); c_rem.alignment = left_align()
            mb(row, 4, row + len(AUTOMATED_KEY_FEATURES) - 1, 6)
        row += 1
    row += 2

    # ══ INBUILT FEATURES ══
    mb(row, 1, row, 6)
    c = ws.cell(row=row, column=1, value="Inbuilt features")
    c.font = section_font(); c.alignment = center_align()
    c.border = medium_border(); c.fill = SECTION_FILL
    ws.row_dimensions[row].height = 20
    row += 1

    for ci, h in enumerate(["Sr.no", "Description", "Vendor Scope (Yes/No)", "Remarks", "", ""], 1):
        c = ws.cell(row=row, column=ci, value=h if h else "")
        if h: style_header_cell(c)
        else: c.border = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    for i, feat in enumerate(AUTOMATED_INBUILT, 1):
        ws.row_dimensions[row].height = 18
        c1 = ws.cell(row=row, column=1, value=i)
        c1.font = normal_font(); c1.alignment = center_align(); c1.border = thin_border()
        c2 = ws.cell(row=row, column=2, value=feat)
        c2.font = normal_font(); c2.alignment = left_align(); c2.border = thin_border()
        ib_row = df_ib[df_ib["Description"] == feat] if df_ib is not None else pd.DataFrame()
        scope_v = str(ib_row.iloc[0].get("Vendor Scope (Yes/No)", "")) if not ib_row.empty else ""
        c3 = ws.cell(row=row, column=3, value=scope_v)
        c3.font = normal_font(); c3.alignment = center_align(); c3.border = thin_border()
        rem_v = ""
        if i == 1: rem_v = "All these key features including at Vendor Side."
        if i == 1:
            mb(row, 4, row + len(AUTOMATED_INBUILT) - 1, 6)
        c4 = ws.cell(row=row, column=4, value=rem_v)
        c4.font = normal_font(); c4.alignment = left_align(); c4.border = thin_border()
        for ci in [5, 6]:
            ws.cell(row=row, column=ci).border = thin_border()
        row += 1
    row += 2

    # ══ INSTALLATION ACCOUNTABILITY ══
    mb(row, 1, row, 6)
    c = ws.cell(row=row, column=1, value="Installation Accountability")
    c.font = section_font(); c.alignment = center_align()
    c.border = medium_border(); c.fill = SECTION_FILL
    ws.row_dimensions[row].height = 20
    row += 1

    for ci, h in enumerate(["Sr.no", "Category",
                             "Vendor Scope (Yes/No)", "Customer Scope\n(Yes/No)", "Remarks", ""], 1):
        c = ws.cell(row=row, column=ci, value=h)
        if h: style_header_cell(c)
        else: c.border = thin_border()
    ws.row_dimensions[row].height = 30
    row += 1

    for i, cat in enumerate(AUTOMATED_INSTALLATION, 1):
        ws.row_dimensions[row].height = 18
        ia_r = df_ia[df_ia["Category"] == cat] if df_ia is not None else pd.DataFrame()
        vendor_s = str(ia_r.iloc[0].get("Vendor Scope (Yes/No)", "")) if not ia_r.empty else ""
        cust_s   = str(ia_r.iloc[0].get("Customer Scope (Yes/No)", "")) if not ia_r.empty else ""
        rem_s    = str(ia_r.iloc[0].get("Remarks", "")) if not ia_r.empty else ""
        vals = [i, cat, vendor_s, cust_s, rem_s, ""]
        aligns = ["center", "left", "center", "center", "left", "center"]
        for ci, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = normal_font()
            c.alignment = center_align() if a == "center" else left_align()
            c.border = thin_border()
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  EXCEL EXPORT — Storage Container
#  A4 landscape, 21 columns incl. image placeholder
# ─────────────────────────────────────────────
def export_container_excel(df, images_dict, project_name, company):
    wb = Workbook()
    ws = wb.active
    ws.title = "Storage Containers"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    headers = [
        "Sr.No", "Description", "Material\nType",
        "Outer\nLength", "Outer\nWidth", "Outer\nHeight",
        "Inner\nLength", "Inner\nWidth", "Inner\nHeight",
        "UOM", "Base\nType", "Colour",
        "Weight\n(Kg)", "Load\nCapacity", "Stackable",
        "Bin Cover\n/Open", "Rate", "Qty",
        "Conceptual\nImage", "Remarks",
    ]
    col_widths = [
        6, 18, 12,
        9, 9, 9,
        9, 9, 9,
        9, 11, 10,
        10, 12, 10,
        12, 9, 7,
        22, 18,
    ]
    IMG_COL_IDX = 19   # 1-based: column 19 = "Conceptual Image"
    IMG_ROW_H   = 55   # row height in points for image rows

    # Title
    n_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value="REQUEST FOR QUOTATION — Storage Containers")
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.alignment = center_align(); t.fill = TITLE_FILL
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    t2 = ws.cell(row=2, column=1,
                 value=f"Project: {project_name}  |  Company: {company}")
    t2.font = Font(name="Arial", bold=True, size=10)
    t2.alignment = center_align(); t2.fill = SECTION_FILL; t2.border = thin_border()
    ws.row_dimensions[2].height = 18

    # Header row (row 4)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=ci, value=h)
        style_header_cell(c)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 36

    # Data rows
    data_row = 5
    valid_rows = df[df["Description"].astype(str).str.strip() != ""].reset_index(drop=True)

    for idx, row in valid_rows.iterrows():
        ws.row_dimensions[data_row].height = IMG_ROW_H

        vals = [
            idx + 1,
            str(row.get("Description", "")),
            str(row.get("Material Type", "")),
            str(row.get("Outer Length", "")),
            str(row.get("Outer Width", "")),
            str(row.get("Outer Height", "")),
            str(row.get("Inner Length", "")),
            str(row.get("Inner Width", "")),
            str(row.get("Inner Height", "")),
            str(row.get("UOM", "")),
            str(row.get("Base Type", "")),
            str(row.get("Colour", "")),
            str(row.get("Weight Kg", "")),
            str(row.get("Load Capacity", "")),
            str(row.get("Stackable", "")),
            str(row.get("Bin Cover/Open", "")),
            str(row.get("Rate", "")),
            str(row.get("Qty", "")),
            "",   # image placeholder
            str(row.get("Remarks", "")),
        ]
        aligns = ["center","left","center","center","center","center",
                  "center","center","center","center","center","center",
                  "center","center","center","center","center","center",
                  "center","left"]

        for ci, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=data_row, column=ci, value=v)
            style_data_cell(c, a)

        # Embed image if present
        img_bytes = images_dict.get(idx)
        if isinstance(img_bytes, bytes):
            try:
                pil_img = PILImage.open(io.BytesIO(img_bytes)).convert("RGB")
                # scale to fit cell
                max_w, max_h = 140, 68
                pil_img.thumbnail((max_w, max_h), PILImage.LANCZOS)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                    pil_img.save(tmp.name, "PNG")
                    xl_img = XLImage(tmp.name)
                    xl_img.width  = pil_img.width
                    xl_img.height = pil_img.height
                    col_letter = get_column_letter(IMG_COL_IDX)
                    ws.add_image(xl_img, f"{col_letter}{data_row}")
                os.remove(tmp.name)
            except Exception:
                pass

        data_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  STREAMLIT UI
# ─────────────────────────────────────────────
st.title("🏭 Warehouse RFQ Generator")
st.markdown("---")

# ── Project info ──
with st.expander("📋 Project Details", expanded=True):
    c1, c2 = st.columns(2)
    project_name = c1.text_input("Project / RFQ Name*", placeholder="e.g. Warehouse Phase 2")
    company_name = c2.text_input("Company Name*", placeholder="e.g. Pinnacle Mobility Pvt. Ltd.")

st.markdown("---")

# ── Category selector ──
st.subheader("Select Item Category")
category = st.selectbox(
    "Item Category",
    options=CATEGORY_OPTIONS,
    index=0,
    help="Choose the category — fields will change automatically"
)

st.markdown("---")

# ══════════════════════════════════════════════
#  STORAGE SYSTEM / MATERIAL HANDLING / DOCK LEVELLER
#  → Sr.no | Item Name (dropdown) | Description | Qty | Unit | Remarks
# ══════════════════════════════════════════════
if category in ("Storage System", "Material Handling", "Dock Leveller"):
    item_options = [""] + CATEGORY_ITEMS[category]
    state_key = f"rfq_df_{category}"

    if state_key not in st.session_state:
        st.session_state[state_key] = pd.DataFrame([{
            "Item Name": "", "Description / Specification": "",
            "Quantity": 1, "Unit": "Nos", "Remarks": ""
        }])

    with st.expander(f"📋 {category} — Item List", expanded=True):
        st.info("Select item from the dropdown, fill description, qty and unit. Use ➕ at the bottom to add more rows.")
        df_edit = st.session_state[state_key].copy()
        df_edit["Quantity"] = pd.to_numeric(df_edit.get("Quantity", 1), errors="coerce").fillna(1).astype(int)
        for col in ["Item Name", "Description / Specification", "Unit", "Remarks"]:
            if col not in df_edit.columns: df_edit[col] = ""
            df_edit[col] = df_edit[col].astype(str).replace("nan", "")

        edited_df = st.data_editor(
            df_edit, num_rows="dynamic", use_container_width=True,
            column_config={
                "Item Name": st.column_config.SelectboxColumn(
                    "Item Name", width="medium", options=item_options),
                "Description / Specification": st.column_config.TextColumn(
                    "Description / Specification", width="large"),
                "Quantity": st.column_config.NumberColumn(
                    "Qty", width="small", min_value=0, step=1, format="%d"),
                "Unit": st.column_config.SelectboxColumn(
                    "Unit", width="small", options=UNIT_OPTIONS),
                "Remarks": st.column_config.TextColumn("Remarks", width="medium"),
            }, key=f"editor_{category}"
        )
        st.session_state[state_key] = edited_df
        valid = edited_df[edited_df["Item Name"].astype(str).str.strip() != ""]
        if len(valid): st.success(f"✅ {len(valid)} item(s) added")

    if st.button("📥 Generate & Download RFQ Excel", type="primary", use_container_width=True):
        if not project_name or not company_name:
            st.error("Please fill in Project Name and Company Name.")
        elif valid.empty:
            st.error("Please add at least one item.")
        else:
            with st.spinner("Generating Excel..."):
                buf = export_standard_excel(category, valid, project_name, company_name)
            fname = f"RFQ_{category.replace(' ', '_')}_{project_name.replace(' ', '_')}.xlsx"
            st.download_button("⬇️ Download Excel", data=buf, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)


# ══════════════════════════════════════════════
#  AUTOMATED STORAGE SYSTEM
#  → Item selector + full spec tables from Excel template
# ══════════════════════════════════════════════
elif category == "Automated Storage System":
    item_options = [""] + AUTOMATED_STORAGE_ITEMS

    with st.expander("🎠 Automated Storage System — Items", expanded=True):
        st.info("Select the carousel/automated storage items you need.")
        auto_state = "rfq_df_Automated Storage System"
        if auto_state not in st.session_state:
            st.session_state[auto_state] = pd.DataFrame([{
                "Item Name": "", "Description / Specification": "",
                "Quantity": 1, "Unit": "Nos", "Remarks": ""
            }])
        auto_df_edit = st.session_state[auto_state].copy()
        auto_df_edit["Quantity"] = pd.to_numeric(auto_df_edit.get("Quantity", 1), errors="coerce").fillna(1).astype(int)
        for col in ["Item Name", "Description / Specification", "Unit", "Remarks"]:
            if col not in auto_df_edit.columns: auto_df_edit[col] = ""
            auto_df_edit[col] = auto_df_edit[col].astype(str).replace("nan", "")

        auto_edited = st.data_editor(
            auto_df_edit, num_rows="dynamic", use_container_width=True,
            column_config={
                "Item Name": st.column_config.SelectboxColumn(
                    "Item Name", width="medium", options=item_options),
                "Description / Specification": st.column_config.TextColumn(
                    "Description / Specification", width="large"),
                "Quantity": st.column_config.NumberColumn(
                    "Qty", width="small", min_value=0, step=1, format="%d"),
                "Unit": st.column_config.SelectboxColumn(
                    "Unit", width="small", options=UNIT_OPTIONS),
                "Remarks": st.column_config.TextColumn("Remarks", width="medium"),
            }, key="editor_automated"
        )
        st.session_state[auto_state] = auto_edited
        auto_valid = auto_edited[auto_edited["Item Name"].astype(str).str.strip() != ""]
        if len(auto_valid): st.success(f"✅ {len(auto_valid)} item(s) selected")

    st.markdown("---")

    model_header = st.text_input(
        "Model Header",
        value="3400 (L) x 3200 (W)   -  465 kgs/tray  -  28 m Height",
        key="model_header_input"
    )

    with st.expander("📐 Model Details (pre-filled — edit Requirement column)", expanded=False):
        md_df = pd.DataFrame(AUTOMATED_MODEL_DETAILS,
                             columns=["Sr", "Category", "Description", "Unit", "Requirement"])
        md_df = md_df.astype(str).replace("nan", "")
        edited_md = st.data_editor(md_df, num_rows="fixed", use_container_width=True,
            column_config={
                "Sr":          st.column_config.TextColumn("Sr.", width="small"),
                "Category":    st.column_config.TextColumn("Category", width="medium"),
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Unit":        st.column_config.TextColumn("Unit", width="small"),
                "Requirement": st.column_config.TextColumn("Requirement ✏️", width="medium"),
            }, key="md_editor")

    with st.expander("⭐ Key Features (fill Status column)", expanded=False):
        kf_df = pd.DataFrame([{"Description": d, "Status": "", "Remarks": ""} for d in AUTOMATED_KEY_FEATURES])
        if "kf_df_state" not in st.session_state: st.session_state["kf_df_state"] = kf_df
        edited_kf = st.data_editor(st.session_state["kf_df_state"], num_rows="fixed",
            use_container_width=True,
            column_config={
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Status":      st.column_config.TextColumn("Status ✏️", width="small"),
                "Remarks":     st.column_config.TextColumn("Remarks ✏️", width="large"),
            }, key="kf_editor")
        st.session_state["kf_df_state"] = edited_kf

    with st.expander("🔧 Inbuilt Features (fill Vendor Scope column)", expanded=False):
        ib_df = pd.DataFrame([{"Description": d, "Vendor Scope (Yes/No)": "", "Remarks": ""}
                               for d in AUTOMATED_INBUILT])
        if "ib_df_state" not in st.session_state: st.session_state["ib_df_state"] = ib_df
        edited_ib = st.data_editor(st.session_state["ib_df_state"], num_rows="fixed",
            use_container_width=True,
            column_config={
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Vendor Scope (Yes/No)": st.column_config.SelectboxColumn(
                    "Vendor Scope ✏️", width="small", options=["", "Yes", "No"]),
                "Remarks": st.column_config.TextColumn("Remarks ✏️", width="large"),
            }, key="ib_editor")
        st.session_state["ib_df_state"] = edited_ib

    with st.expander("🏗️ Installation Accountability (fill Scope columns)", expanded=False):
        ia_df = pd.DataFrame([{
            "Category": c, "Vendor Scope (Yes/No)": "",
            "Customer Scope (Yes/No)": "", "Remarks": ""
        } for c in AUTOMATED_INSTALLATION])
        if "ia_df_state" not in st.session_state: st.session_state["ia_df_state"] = ia_df
        edited_ia = st.data_editor(st.session_state["ia_df_state"], num_rows="fixed",
            use_container_width=True,
            column_config={
                "Category": st.column_config.TextColumn("Category", width="large"),
                "Vendor Scope (Yes/No)": st.column_config.SelectboxColumn(
                    "Vendor Scope ✏️", width="small", options=["", "Yes", "No"]),
                "Customer Scope (Yes/No)": st.column_config.SelectboxColumn(
                    "Customer Scope ✏️", width="small", options=["", "Yes", "No"]),
                "Remarks": st.column_config.TextColumn("Remarks ✏️", width="medium"),
            }, key="ia_editor")
        st.session_state["ia_df_state"] = edited_ia

    if st.button("📥 Generate & Download RFQ Excel", type="primary", use_container_width=True):
        if not project_name or not company_name:
            st.error("Please fill in Project Name and Company Name.")
        else:
            with st.spinner("Generating Excel..."):
                buf = export_automated_excel(
                    edited_md, model_header,
                    edited_kf, edited_ib, edited_ia,
                    project_name, company_name
                )
            fname = f"RFQ_Automated_Storage_{project_name.replace(' ', '_')}.xlsx"
            st.download_button("⬇️ Download Excel", data=buf, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)


# ══════════════════════════════════════════════
#  STORAGE CONTAINER
#  → Full 20-column table + image upload
# ══════════════════════════════════════════════
elif category == "Storage Container":
    container_options = [""] + STORAGE_CONTAINER_ITEMS

    if "sc_df" not in st.session_state:
        st.session_state["sc_df"] = pd.DataFrame([{
            "Description": "", "Material Type": "",
            "Outer Length": "", "Outer Width": "", "Outer Height": "",
            "Inner Length": "", "Inner Width": "", "Inner Height": "",
            "UOM": "", "Base Type": "", "Colour": "",
            "Weight Kg": "", "Load Capacity": "",
            "Stackable": "", "Bin Cover/Open": "",
            "Rate": "", "Qty": 1, "Remarks": ""
        }])
    if "sc_images" not in st.session_state:
        st.session_state["sc_images"] = {}

    with st.expander("📦 Storage Container — Item Details", expanded=True):
        st.info("Select container type from dropdown, fill all dimensions. Upload conceptual images on the right.")

        sc_df_edit = st.session_state["sc_df"].copy()
        sc_df_edit["Qty"] = pd.to_numeric(sc_df_edit.get("Qty", 1), errors="coerce").fillna(1).astype(int)
        for col in ["Description","Material Type","Outer Length","Outer Width","Outer Height",
                    "Inner Length","Inner Width","Inner Height","UOM","Base Type","Colour",
                    "Weight Kg","Load Capacity","Stackable","Bin Cover/Open","Rate","Remarks"]:
            if col not in sc_df_edit.columns: sc_df_edit[col] = ""
            sc_df_edit[col] = sc_df_edit[col].astype(str).replace("nan","")

        ed_col, img_col = st.columns([3, 1])
        with ed_col:
            edited_sc = st.data_editor(
                sc_df_edit, num_rows="dynamic", use_container_width=True,
                column_config={
                    "Description":   st.column_config.SelectboxColumn(
                        "Container Type ▼", width="medium", options=container_options),
                    "Material Type": st.column_config.SelectboxColumn(
                        "Material ▼", width="small",
                        options=["","Plastic","Metal","Wood","Corrugated","Fibre","Other"]),
                    "Outer Length":  st.column_config.TextColumn("Outer L (mm)", width="small"),
                    "Outer Width":   st.column_config.TextColumn("Outer W (mm)", width="small"),
                    "Outer Height":  st.column_config.TextColumn("Outer H (mm)", width="small"),
                    "Inner Length":  st.column_config.TextColumn("Inner L (mm)", width="small"),
                    "Inner Width":   st.column_config.TextColumn("Inner W (mm)", width="small"),
                    "Inner Height":  st.column_config.TextColumn("Inner H (mm)", width="small"),
                    "UOM":           st.column_config.SelectboxColumn(
                        "UOM ▼", width="small", options=[""] + UNIT_OPTIONS),
                    "Base Type":     st.column_config.SelectboxColumn(
                        "Base Type ▼", width="small",
                        options=["","Flat","Ribbed","Louvred","Grid","Other"]),
                    "Colour":        st.column_config.TextColumn("Colour", width="small"),
                    "Weight Kg":     st.column_config.TextColumn("Weight (Kg)", width="small"),
                    "Load Capacity": st.column_config.TextColumn("Load Cap (Kg)", width="small"),
                    "Stackable":     st.column_config.SelectboxColumn(
                        "Stackable ▼", width="small", options=["","Yes","No","N/A"]),
                    "Bin Cover/Open":st.column_config.SelectboxColumn(
                        "Cover/Open ▼", width="small",
                        options=["","Open","Covered","Lid","N/A"]),
                    "Rate":          st.column_config.TextColumn("Rate", width="small"),
                    "Qty":           st.column_config.NumberColumn(
                        "Qty", width="small", min_value=0, step=1),
                    "Remarks":       st.column_config.TextColumn("Remarks", width="medium"),
                }, key="sc_editor"
            )
        st.session_state["sc_df"] = edited_sc

        with img_col:
            st.write("**Conceptual Images**")
            for i in range(len(edited_sc)):
                desc = str(edited_sc.iloc[i].get("Description","")).strip()
                label = f"Row {i+1}: {desc}" if desc else f"Row {i+1}"
                f = st.file_uploader(label, type=["png","jpg","jpeg"],
                                     key=f"sc_img_{i}")
                if f:
                    st.session_state["sc_images"][i] = f.getvalue()
                    st.image(f.getvalue(), width=80)

        valid_sc = edited_sc[edited_sc["Description"].astype(str).str.strip() != ""]
        if len(valid_sc): st.success(f"✅ {len(valid_sc)} container type(s) defined")

    if st.button("📥 Generate & Download RFQ Excel", type="primary", use_container_width=True):
        if not project_name or not company_name:
            st.error("Please fill in Project Name and Company Name.")
        elif valid_sc.empty:
            st.error("Please define at least one container type.")
        else:
            with st.spinner("Generating Excel..."):
                buf = export_container_excel(
                    valid_sc.reset_index(drop=True),
                    st.session_state["sc_images"],
                    project_name, company_name
                )
            fname = f"RFQ_Storage_Container_{project_name.replace(' ','_')}.xlsx"
            st.download_button("⬇️ Download Excel", data=buf, file_name=fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
